from datetime import datetime
import json
import os
import re
import time
import subprocess
from bs4 import BeautifulSoup
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook

# Crear carpetas si no existen y asegurarse de que tienen los permisos adecuados
def crear_carpetas():
    for carpeta in ['HTMLMobile', 'HTMLDesktop']:
        if not os.path.exists(carpeta):
            os.makedirs(carpeta)
        os.chmod(carpeta, 0o777)

# Función para verificar si una URL responde con un código de estado 200 y saltar pagina 404 de Entel
def validar_Url(url):
    try:
        response = requests.head(url, allow_redirects=True)
        codigo = response.status_code
        if codigo == 200:
            response = requests.get(url)
            soup = BeautifulSoup(response.content, 'html.parser')
            titulo = soup.title.string if soup.title else ""

            if "Error" in titulo or "Página de Error | Entel" in titulo:
                codigo = 404
                descripcion = "Redireccion - Página de Error Detectada"
                print(f"Error al verificar la URL {url}: {codigo} {descripcion}")
            else:
                descripcion = "OK"
        else:
            descripcion = requests.status_codes._codes[codigo][0].replace('_', ' ').title()

        return codigo, descripcion
    except requests.RequestException as e:
        print(f"Error al verificar la URL {url}: {e}")
        return None, str(e)

# Función para ejecutar Lighthouse en una URL específica (Mobile - Desktop)
def auditoria_Lighthouse(url, mode):
    nombreLimpio = re.sub(r'[^\w.-]', '_', url)
    if mode == 'mobile':
        finalHTML = os.path.join('HTMLMobile', f'{mode}_{nombreLimpio}.html')
    else:
        finalHTML = os.path.join('HTMLDesktop', f'{mode}_{nombreLimpio}.html')

    #username = os.getlogin()
    
    # Ruta completa al ejecutable de Node.js
    #node_path = f'/Users/{username}/.nvm/versions/node/v20.15.1/bin/node'
    PATH_NODE = '/usr/bin/node'
    #PATH_NODE = r'C:\Program Files\nodejs\node.exe'

    # Ruta completa al archivo de Lighthouse
    #lighthouse_path = f'/Users/{username}/.nvm/versions/node/v20.15.1/lib/node_modules/lighthouse/cli/index.js'
    LIGHTHOUSE_PATH = '/usr/lib/node_modules/lighthouse/cli/index.js'
    #LIGHTHOUSE_PATH = rf'C:\Users\{username}\AppData\Roaming\npm\node_modules\lighthouse\cli\index.js'

    # Comando para ejecutar Lighthouse con la configuración necesaria
    command = [
        PATH_NODE,
        LIGHTHOUSE_PATH,
        url,
        '--output=html',
        f'--output-path={finalHTML}',
        '--chrome-flags="--headless --no-sandbox --disable-dev-shm-usage --window-size=1920,1080"'
    ]

    # Configuración extra para el modo Desktop
    if mode == 'desktop':
        command.append('--preset=desktop')

    try:
        # Ejecuta la auditoría de Lighthouse
        result = subprocess.run(command, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"Error al ejecutar Lighthouse desde {url} ({mode}):\n{result.stderr}")
            return None
    except FileNotFoundError:
        print(f"Lighthouse no se encontró en la ruta especificada. Asegúrate de que está instalado y accesible en {LIGHTHOUSE_PATH}.")
        return None
    
    print(f"Se genera informe {finalHTML}")
    return finalHTML

# Función que ejecuta Lighthouse para una URL
def urls_Lighthouse(url):
    codigo, descripcionCodigo = validar_Url(url)
    if codigo is None or codigo != 200:
        # Si la URL no es válida, registrar el error en el Excel
        actualizar_Excel(url, {'performance': None, 'accessibility': None, 'seo': None}, {'performance': None, 'accessibility': None, 'seo': None}, "Error", descripcionCodigo)
        return {
            'url': url,
            'totalTest': None
        }

    inicio = time.time()

    # Ejecutar Lighthouse (Mobile - Desktop)
    print(f"Ejecutando auditoria Lighthouse ...")
    rporteMOBILE = auditoria_Lighthouse(url, 'mobile')
    if rporteMOBILE:
        puntuacionesMOBILE = extraer_Puntuaciones(rporteMOBILE)
    else:
        puntuacionesMOBILE = {'performance': None, 'accessibility': None, 'seo': None}

    reporteDESKTOP = auditoria_Lighthouse(url, 'desktop')
    if reporteDESKTOP:
        puntuacionesDESKTOP = extraer_Puntuaciones(reporteDESKTOP)
    else:
        puntuacionesDESKTOP = {'performance': None, 'accessibility': None, 'seo': None}

    # Actualizar el archivo Excel después de completar ambas auditorías
    actualizar_Excel(url, puntuacionesMOBILE, puntuacionesDESKTOP, codigo, descripcionCodigo)
    
    termino = time.time()
    totalTest = round(termino - inicio, 2)
    
    return {
        'url': url,
        'totalTest': totalTest
    }

# Función para extraer las puntuaciones de SEO, Accesibilidad y Performance del reporte HTML
def extraer_Puntuaciones(pathHTML):
    try:
        puntuaciones = {
            'performance': None,
            'accessibility': None,
            'seo': None
        }

        with open(pathHTML, 'r', encoding="utf-8") as file:
            filas = file.readlines()
        
        # Buscamos la línea que contiene los datos JSON
        jsonLine = next((line for line in filas if 'window.__LIGHTHOUSE_JSON__' in line), None)
        if jsonLine:
            match = re.search(r'=(.*?);<', jsonLine)
            if match:
                resultados = match.group(1)
            else:
                resultados = None
            data = json.loads(resultados)
            puntuaciones['performance'] = int(data["categories"]["performance"]["score"]*100) if data["categories"]["performance"]["score"] is not None else "N/A"
            puntuaciones['accessibility'] = int(data["categories"]["accessibility"]["score"]*100) if data["categories"]["accessibility"]["score"] is not None else "N/A"
            puntuaciones['seo'] = int(data["categories"]["seo"]["score"]*100) if data["categories"]["seo"]["score"] is not None else "N/A"

        return puntuaciones
    except Exception as e:
        print(f"Error al extraer puntuaciones desde {pathHTML}: {e}")
        return {
            'performance': None,
            'accessibility': None,
            'seo': None
        }

# Función para crear y actualizar el archivo Excel con los resultados
# Función para crear y actualizar el archivo Excel con los resultados
def actualizar_Excel(url, puntuacionesMOBILE, puntuacionesDESKTOP, codigo, descripcionCodigo):
    global pathArchivo
    fecha_hora_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Nombre del archivo con la fecha y hora de ejecución (solo se crea una vez)
    if not pathArchivo:
        pathArchivo = f'resultados_{fecha_hora_actual}.xlsx'
    
    try:
        cargarExcel = load_workbook(pathArchivo)
        hoja = cargarExcel.active
    except FileNotFoundError:
        cargarExcel = Workbook()
        hoja = cargarExcel.active
        hoja.append(['URL', 'Performance Mobile', 'Performance Desktop', 'Accesibilidad Mobile', 'Accesibilidad Desktop', 'SEO Mobile', 'SEO Desktop', 'Código', 'Descripción Código'])

    actualizado = False
    for row in hoja.iter_rows(min_row=2, values_only=False):
        if row[0].value == url:
            row[1].value = puntuacionesMOBILE['performance']
            row[2].value = puntuacionesDESKTOP['performance']
            row[3].value = puntuacionesMOBILE['accessibility']
            row[4].value = puntuacionesDESKTOP['accessibility']
            row[5].value = puntuacionesMOBILE['seo']
            row[6].value = puntuacionesDESKTOP['seo']
            row[7].value = codigo
            row[8].value = descripcionCodigo
            actualizado = True
            break

    if not actualizado:
        new_row = [
            url,
            puntuacionesMOBILE['performance'], puntuacionesDESKTOP['performance'],
            puntuacionesMOBILE['accessibility'], puntuacionesDESKTOP['accessibility'],
            puntuacionesMOBILE['seo'], puntuacionesDESKTOP['seo'],
            codigo, descripcionCodigo
        ]
        hoja.append(new_row)

    # Ajustar el ancho de las columnas
    for col in hoja.columns:
        largo = 0
        columna = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > largo:
                    largo = len(cell.value)
            except:
                pass
        ancho = (largo + 2)
        hoja.column_dimensions[columna].width = ancho

    cargarExcel.save(pathArchivo)

# Crear carpetas si no existen y asegurarse de que tienen los permisos adecuados
crear_carpetas()

# Leer las URLs desde un archivo de texto
with open('urls.txt', 'r') as archivo:
    urls = archivo.read().splitlines()

# Variable global para almacenar la ruta del archivo Excel
pathArchivo = None

with ThreadPoolExecutor(max_workers=1) as ejec:
    future_to_url = {ejec.submit(urls_Lighthouse, url): url for url in urls}
    
    for future in as_completed(future_to_url):
        url = future_to_url[future]
        try:
            result = future.result()
            if result['totalTest'] is not None:
                print(f"Duración total de la auditoría para {url}: {result['totalTest']} segundos")
        except Exception as e:
            print(f"Error de procesamiento en {url}: {e}")
    
